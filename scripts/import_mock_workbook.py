# -*- coding: utf-8 -*-
"""Convert the client's updated mock workbook into literal SQL seed data.

Emits INSERTs for tenants/projects/survey_responses/survey_answers that mirror
the three sheets exactly. Sentiment is heuristic-derived (workbook has none).
"""
import openpyxl, uuid, re, sys

WB = r"d:\Work\Projects\Upwork\Data Monitoring Platform (Base 44 Dashboard)\Resources\Mock Data_Field Survey - Online Update_Rev 4.xlsx"
OUT = r"C:\Users\GHOST\AppData\Local\Temp\claude\d--Work-Projects-Upwork-Data-Monitoring-Platform--Base-44-Dashboard-\ef16fa0d-1381-4c9d-8b1d-cc5eaca37ea8\scratchpad\literal_seed_fragment.sql"

TENANT = "cln-0010"

def clean(s):
    """Fix mojibake: digit�digit -> '-', ' �NN' -> ' £NN', stray � -> "'"."""
    if s is None: return None
    s = str(s).strip()
    s = re.sub(r"(?<=\d)�(?=\d)", "-", s)
    s = re.sub(r"�(?=\d)", "£", s)
    s = s.replace("�", "'")
    # normalize unicode dashes in brackets like 25–34
    s = re.sub(r"(?<=\d)[–—](?=\d)", "-", s)
    return s

def esc(s):
    return s.replace("'", "''")

NEG = ["worse","unsafe","poor","unclear","slow","doubled","tripled","uncapped",
       "overflow","dust","noise","grime","tough","no choice","zero","blindly",
       "out of nowhere","won't","wont","complaint","never","frustrat","annoy",
       "problem","issue","fail","broke","dirty","dark","scary","crime","expensive",
       "messy","cramped","lacks","desperately","complicated","impossible","block"]
POS = ["great","love","happy","improved","brilliant","fantastic","good","thank",
       "brightened","amazing","excellent","appreciate","keen","enjoy",
       "wonderful","pleased","impress"]

# Hand overrides where the lexicon misreads a text (demo-balanced tagging:
# community-supportive suggestions read positive; retrospective complaints
# read negative). Keep in sync with supabase/_patch_sentiment.sql.
OVERRIDES = {
    "Please add a community repair workshop": "neutral",
    "Could you install a pet-wash station": "positive",
    # neutral -> positive
    "Please partner with the local bakery": "positive",
    "Can the development host a monthly": "positive",
    "communal lounge host free wellbeing": "positive",
    "share regular updates on local businesses": "positive",
    "lease add motion-sensor LED lighting": "positive",
    "A communal rooftop garden with BBQ": "positive",
    "adding a small, fenced dog run": "positive",
    "An 8-storey building needs dedicated kid": "positive",
    "Planting mature, dense canopy trees": "positive",
    "Instead of a generic concrete plaza": "positive",
    "More construciton site jobs please": "positive",
    "Will the new building include public amenities": "positive",
    "Please prioritize hiring local subcontractors": "positive",
    "Can we get weekly text or email updates": "positive",
    # neutral -> negative
    "Please respect the agreed-upon working hours": "negative",
    "Can you provide a clear timeline": "negative",
    "We need a clear, public point of contact": "negative",
    "Development should have included dedicated": "negative",
    "Acoustic insulation between the gym": "negative",
    "installing a third high-speed service elevator": "negative",
}

def sentiment(text):
    for prefix, s in OVERRIDES.items():
        if text.startswith(prefix):
            return s, {"positive": 0.5, "neutral": 0.0, "negative": -0.5}[s]
    t = text.lower()
    n = sum(1 for w in NEG if w in t)
    p = sum(1 for w in POS if w in t)
    if n > p: return "negative", -0.5
    if p > n: return "positive", 0.5
    return "neutral", 0.0

def bracket(age):
    a = float(age)
    if a < 25: return "18-24"
    if a < 35: return "25-34"
    if a < 50: return "35-49"
    if a < 65: return "50-64"
    return "65+"

def rid(tag):
    return str(uuid.uuid5(uuid.NAMESPACE_URL, "mockdata/" + tag))

def norm15(v):
    return (float(v) - 1) / 4 * 100

def numish(v):
    """Int, or None for blanks / 'N/A' (Rev 4 has one N/A cost cell)."""
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def ts(i):
    return f"2026-07-{1 + (i % 5):02d} {9 + (i % 8):02d}:{(i * 7) % 60:02d}:00+00"

wb = openpyxl.load_workbook(WB, data_only=True)
sql = []
sql.append("-- ==== LITERAL MOCK DATA (generated from Mock Data_Field Survey - Online Survey_Update.xlsx) ====")
sql.append(f"insert into public.tenants (id, name, branding) values")
sql.append(f"  ('{TENANT}', 'Client CLN-0010', '{{\"brand\":\"37 99 235\",\"logoText\":\"CLN-0010\"}}');")
sql.append("")
sql.append("insert into public.projects (id, tenant_id, name, status, completion_date, retention_expires_at) values")
sql.append(f"  ('{rid('proj-AST-0001')}', '{TENANT}', 'AST-0001 (In-Construction)', 'active', null, null),")
sql.append(f"  ('{rid('proj-AST-0004')}', '{TENANT}', 'AST-0004 (Completed - BTR)', 'in_report', date '2026-05-01', date '2026-11-01'),")
sql.append(f"  ('{rid('proj-AST-0006')}', '{TENANT}', 'AST-0006 (Completed - Private Sale)', 'in_report', date '2026-04-15', date '2026-10-15');")
sql.append("")

sent_log = []

def answer(resp, code, raw, rtype, vnum=None, vnorm=None, sent=None):
    sent_sql = "'" + sent + "'" if sent else "null"
    vnum_sql = str(vnum) if vnum is not None else "null"
    vnorm_sql = str(vnorm) if vnorm is not None else "null"
    return (f"('{resp}','{code}','{esc(str(raw))}','{rtype}',"
            f"{vnum_sql},{vnorm_sql},{sent_sql})")

answers_rows = []
resp_rows = []

# ---------- Field Survey ----------
ws = wb["Field Survey"]
rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
for i, r in enumerate(rows[1:]):  # skip header
    tag = f"field-{i}"
    respid = rid(tag)
    age = r[4]; access = r[5]; prox = clean(r[6]); q4 = r[7]; q5 = r[8]
    well = clean(r[9]); chip = clean(r[10]); text = clean(r[11]) or ""
    text = text[:280]
    sen, sscore = sentiment(text)
    sent_log.append(("FIELD", text[:70], sen))
    cohort = clean(r[3])
    resp_rows.append(
      f"('{respid}','{TENANT}','{rid('proj-AST-0001')}','field','field_pwa','in_construction',null,"
      f"'construction_adjacent',null,'{cohort}',2026,3,timestamptz '{ts(i)}',"
      f"'{bracket(age)}','In-Construction','—',null,'{esc(text)}','{sen}',{sscore})")
    answers_rows.append(answer(respid, "FS_AGE", int(age), "numeric", vnum=int(age)))
    answers_rows.append(answer(respid, "FS_ACCESS_COHORT", int(access), "categorical"))
    answers_rows.append(answer(respid, "FS_PROXIMITY", prox, "categorical"))
    answers_rows.append(answer(respid, "FS_PUBLIC_SPACE", int(q4), "numeric", vnum=int(q4), vnorm=norm15(q4)))
    answers_rows.append(answer(respid, "FS_GRIEVANCE", int(q5), "numeric", vnum=int(q5), vnorm=norm15(q5)))
    wnorm = {"Yes_POS": 100, "YES_NEG": 50, "NO_NEG": 0}.get(well, 0)
    answers_rows.append(answer(respid, "FS_WELLBEING_AWARE", well, "categorical", vnorm=wnorm))
    if chip and chip != "N/A":
        answers_rows.append(answer(respid, "FS_OFFERING", chip, "categorical"))
    answers_rows.append(answer(respid, "FS_OPEN", text, "text", sent=sen))

# ---------- Online sheets ----------
ONLINE = [
    ("Online In Building QR Codes", "AST-0004", "btr", "build_to_rent", "BTR"),
    ("Private Sale Online In Building", "AST-0006", "private_ownership", "build_to_sell", "Private Ownership"),
]
for sheet, asset, tenure, deliv, tenure_label in ONLINE:
    ws = wb[sheet]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    for i, r in enumerate(rows[2:]):  # skip short-header row + question-text row
        tag = f"{asset}-{i}"
        respid = rid(tag)
        agebr = clean(r[5]); occ = clean(r[6]); q3 = r[7]; q3b = clean(r[8])
        q4 = clean(r[9]); q5 = r[10]; q6 = r[11]; q7 = r[12]; q8 = r[13]; q9 = r[14]
        off1 = clean(r[15]); off2 = clean(r[16]); text = clean(r[17]) or ""
        text = text[:280]
        sen, sscore = sentiment(text)
        sent_log.append((sheet[:12], text[:70], sen))
        cohort = clean(r[4])
        resp_rows.append(
          f"('{respid}','{TENANT}','{rid('proj-' + asset)}','online','digital_public','completed','{tenure}',"
          f"'resident_completed','{deliv}','{cohort}',2026,3,timestamptz '{ts(i + 40)}',"
          f"'{agebr}','Completed','{tenure_label}',null,'{esc(text)}','{sen}',{sscore})")
        answers_rows.append(answer(respid, "OL_OCCUPANCY", occ, "categorical"))
        q3n = numish(q3)
        if q3n is not None:
            answers_rows.append(answer(respid, "OL_COST_MANAGEABLE", q3n, "numeric", vnum=q3n, vnorm=norm15(q3n)))
        else:
            answers_rows.append(answer(respid, "OL_COST_MANAGEABLE", clean(q3) or "N/A", "categorical"))
        if q3b:
            q3b = q3b[:280]
            answers_rows.append(answer(respid, "OL_COST_FOLLOWUP", q3b, "text", sent="negative"))
        answers_rows.append(answer(respid, "OL_ENERGY_KNOW", q4, "categorical", vnorm=(100 if str(q4).lower() == "yes" else 0)))
        for code, v in [("OL_ACTIVE_TRAVEL", q5), ("OL_SECURITY", q6), ("OL_PUBLIC_REALM", q7), ("OL_GRIEVANCE", q8), ("OL_WELLBEING_AWARE", q9)]:
            vn = numish(v)
            if vn is not None:
                answers_rows.append(answer(respid, code, vn, "numeric", vnum=vn, vnorm=norm15(vn)))
        answers_rows.append(answer(respid, "OL_OFFERING_1", off1, "multi"))
        answers_rows.append(answer(respid, "OL_OFFERING_2", off2, "multi"))
        answers_rows.append(answer(respid, "OL_OPEN", text, "text", sent=sen))

sql.append("insert into public.survey_responses")
sql.append("  (id, tenant_id, project_id, channel, source, asset_class_state, tenure,")
sql.append("   respondent_typology, delivery_model, temporal_cohort, period_year, period_quarter,")
sql.append("   submitted_at, q1_demographic, q2_asset_class, q3_tenure, housing_cost_to_income,")
sql.append("   q10_text, q10_sentiment, q10_sentiment_score) values")
sql.append(",\n".join(resp_rows) + ";")
sql.append("")
sql.append("insert into public.survey_answers")
sql.append("  (response_id, question_code, value_raw, value_raw_type, value_numeric, value_normalized, sentiment) values")
sql.append(",\n".join(answers_rows) + ";")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql))

print(f"responses: {len(resp_rows)}  answers: {len(answers_rows)}")
print("--- sentiment assignments ---")
for src, t, s in sent_log:
    print(f"[{src:<12}] {s:<8} | {t}")
